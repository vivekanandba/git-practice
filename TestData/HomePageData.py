import openpyxl



class HomePageData:
    test_HomePage_data = [{'firstname':'Vivek', 'lastname':'Anand', 'gender':'Male'},{'firstname':'Asder', 'lastname':'bva', 'gender':'Male'}]

    @staticmethod
    def getTestData(test_case_name):

        book = openpyxl.load_workbook("F:\\Learning\\Programming\\Automation Testing\\Selenium Webdriver with Python from Scratch + Frameworks - Udemy\\selenium-udemy-01\\SeleniumTestFramework\\PythonDemo.xlsx")
        sheet = book.active
        Dict = {}

        for i  in range(1,sheet.max_row+1):
            if sheet.cell(row=i,column=1).value == test_case_name:
        
                for j in range(2,sheet.max_column+1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]